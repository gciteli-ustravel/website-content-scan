import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl

from scripts import update_scan


class UpdateExcelTests(unittest.TestCase):
    def test_update_excel_preserves_missing_rows_and_reports_zero_expired(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            input_path = Path(tmpdir) / "input.xlsx"
            output_path = Path(tmpdir) / "output.xlsx"

            workbook = openpyxl.Workbook()
            ws = workbook.active
            ws.title = update_scan.SHEET_NAME
            ws.append(update_scan.REQUIRED_COLUMNS)
            ws.append(
                [
                    "https://example.com/existing",
                    "example.com",
                    "/existing",
                    "",
                    "In Review",
                    "2024-01-01",
                ]
            )
            ws.append(
                [
                    "https://example.com/keep-me",
                    "example.com",
                    "/keep-me",
                    "",
                    "Published",
                    "2024-02-02",
                ]
            )
            workbook.save(input_path)

            existing_page = update_scan.SitemapPage(
                url="https://example.com/existing",
                key="example.com/existing",
                site="example.com",
                page="/existing",
                sub_page="",
                last_updated="2025-03-03",
            )
            new_page = update_scan.SitemapPage(
                url="https://example.com/new-page",
                key="example.com/new-page",
                site="example.com",
                page="/new-page",
                sub_page="",
                last_updated="2025-04-04",
            )

            summary = update_scan.update_excel(
                input_path,
                output_path,
                {existing_page.key: existing_page, new_page.key: new_page},
                update_scan.DEFAULT_EXPIRED_STATUS,
            )

            self.assertEqual(summary["added"], 1)
            self.assertEqual(summary["updated_last_updated"], 1)
            self.assertEqual(summary["expired"], 0)

            result = openpyxl.load_workbook(output_path)
            ws = result[update_scan.SHEET_NAME]
            self.assertEqual(ws.cell(row=2, column=6).value, "2025-03-03")
            self.assertEqual(ws.cell(row=3, column=5).value, "Published")
            self.assertEqual(ws.cell(row=4, column=1).value, "https://example.com/new-page")
            self.assertEqual(ws.cell(row=4, column=5).value, update_scan.DEFAULT_NEW_STATUS)


class UpdateSmartsheetTests(unittest.TestCase):
    def test_update_smartsheet_preserves_missing_rows_and_reports_zero_expired(self) -> None:
        sheet = {
            "columns": [
                {"id": 1, "title": "URL"},
                {"id": 2, "title": "Site"},
                {"id": 3, "title": "Page"},
                {"id": 4, "title": "Sub-Page"},
                {"id": 5, "title": "Page Status", "type": "TEXT_NUMBER"},
                {"id": 6, "title": "Last Updated"},
            ],
            "rows": [
                {
                    "id": 101,
                    "cells": [
                        {"columnId": 1, "value": "https://example.com/existing"},
                        {"columnId": 5, "value": "In Review"},
                        {"columnId": 6, "value": "2024-01-01"},
                    ],
                },
                {
                    "id": 202,
                    "cells": [
                        {"columnId": 1, "value": "https://example.com/keep-me"},
                        {"columnId": 5, "value": "Published"},
                        {"columnId": 6, "value": "2024-02-02"},
                    ],
                },
            ],
        }

        existing_page = update_scan.SitemapPage(
            url="https://example.com/existing",
            key="example.com/existing",
            site="example.com",
            page="/existing",
            sub_page="",
            last_updated="2025-03-03",
        )
        new_page = update_scan.SitemapPage(
            url="https://example.com/new-page",
            key="example.com/new-page",
            site="example.com",
            page="/new-page",
            sub_page="",
            last_updated="2025-04-04",
        )

        requests_seen = []

        def fake_request(method: str, url: str, token: str, payload=None):
            requests_seen.append((method, url, payload))
            if method == "GET":
                return sheet
            return {}

        with patch.dict(os.environ, {"SMARTSHEET_ACCESS_TOKEN": "token"}, clear=False):
            with patch("scripts.update_scan.smartsheet_request", side_effect=fake_request):
                summary = update_scan.update_smartsheet(
                    "sheet-123",
                    {existing_page.key: existing_page, new_page.key: new_page},
                    update_scan.DEFAULT_EXPIRED_STATUS,
                )

        self.assertEqual(summary["added"], 1)
        self.assertEqual(summary["updated_last_updated"], 1)
        self.assertEqual(summary["expired"], 0)

        put_payloads = [payload for method, _, payload in requests_seen if method == "PUT"]
        post_payloads = [payload for method, _, payload in requests_seen if method == "POST"]

        self.assertEqual(len(put_payloads), 1)
        self.assertEqual(
            put_payloads[0],
            [{"id": 101, "cells": [{"columnId": 6, "value": "2025-03-03"}]}],
        )
        self.assertEqual(len(post_payloads), 1)
        self.assertEqual(post_payloads[0][0]["cells"][0]["value"], "https://example.com/new-page")
        self.assertEqual(post_payloads[0][0]["cells"][4]["value"], update_scan.DEFAULT_NEW_STATUS)


if __name__ == "__main__":
    unittest.main()
