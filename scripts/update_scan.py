#!/usr/bin/env python3
"""Update a Website Content Scan sheet from website sitemaps."""

from __future__ import annotations

import argparse
import datetime as dt
import gzip
import json
import os
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.parse import urlsplit, urlunsplit
from xml.etree import ElementTree

import openpyxl
import requests

try:
    import yaml
except ModuleNotFoundError:  # pragma: no cover - only used on minimally set up machines.
    yaml = None


DEFAULT_CONFIG = Path(__file__).resolve().parents[1] / "sites.yml"
DEFAULT_EXPIRED_STATUS = "Page Expired"
DEFAULT_NEW_STATUS = "Not Started"
SHEET_NAME = "Website Content Scan"
REQUIRED_COLUMNS = ["URL", "Site", "Page", "Sub-Page", "Page Status", "Last Updated"]
SKIP_EXTENSIONS = {
    ".7z",
    ".avi",
    ".bmp",
    ".css",
    ".csv",
    ".doc",
    ".docx",
    ".eot",
    ".gif",
    ".gz",
    ".ico",
    ".jpeg",
    ".jpg",
    ".js",
    ".json",
    ".m4a",
    ".mov",
    ".mp3",
    ".mp4",
    ".ogg",
    ".otf",
    ".pdf",
    ".png",
    ".ppt",
    ".pptx",
    ".rar",
    ".rss",
    ".svg",
    ".tar",
    ".tgz",
    ".tif",
    ".tiff",
    ".ttf",
    ".txt",
    ".wav",
    ".webm",
    ".webp",
    ".woff",
    ".woff2",
    ".xls",
    ".xlsx",
    ".xml",
    ".zip",
}


@dataclass(frozen=True)
class SitemapPage:
    url: str
    key: str
    site: str
    page: str
    sub_page: str
    last_updated: str


def load_config(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8") as handle:
        if yaml:
            config = yaml.safe_load(handle) or {}
        else:
            config = parse_simple_sites_yaml(handle.read())
    sites = config.get("sites")
    if not isinstance(sites, list) or not sites:
        raise ValueError(f"{path} must contain a non-empty 'sites' list")
    for site in sites:
        if not isinstance(site, dict) or not site.get("name") or not site.get("sitemap"):
            raise ValueError("Each site must include 'name' and 'sitemap'")
        exclude_paths = site.get("exclude_paths", [])
        if exclude_paths is None:
            site["exclude_paths"] = []
        elif not isinstance(exclude_paths, list):
            raise ValueError("'exclude_paths' must be a list when provided")
    return sites


def parse_simple_sites_yaml(text: str) -> dict[str, list[dict[str, str]]]:
    sites: list[dict[str, str]] = []
    current: dict[str, str] | None = None
    current_list_key: str | None = None

    for raw_line in text.splitlines():
        line = raw_line.split("#", 1)[0].rstrip()
        stripped = line.strip()
        if not stripped or stripped == "sites:":
            continue
        if stripped.startswith("- ") and line.startswith("  - "):
            if current:
                sites.append(current)
            current = {}
            current_list_key = None
            stripped = stripped[2:].strip()
        elif stripped.startswith("- ") and current is not None and current_list_key:
            current.setdefault(current_list_key, []).append(stripped[2:].strip().strip('"\''))
            continue
        if current is None:
            continue
        if ":" not in stripped:
            continue
        key, value = stripped.split(":", 1)
        key = key.strip()
        value = value.strip().strip('"\'')
        if value:
            current[key] = value
            current_list_key = None
        else:
            current[key] = []
            current_list_key = key

    if current:
        sites.append(current)
    return {"sites": sites}


def strip_www(host: str) -> str:
    host = host.lower()
    return host[4:] if host.startswith("www.") else host


def clean_path(path: str) -> str:
    path = re.sub(r"/+", "/", path or "/")
    if not path.startswith("/"):
        path = f"/{path}"
    if path != "/":
        path = path.rstrip("/")
    return path


def normalize_url(raw_url: str) -> tuple[str, str, str]:
    parsed = urlsplit(str(raw_url).strip())
    if not parsed.netloc:
        raise ValueError(f"Not an absolute URL: {raw_url}")
    host = strip_www(parsed.netloc.split("@")[-1].split(":")[0])
    path = clean_path(parsed.path)
    canonical = urlunsplit(("https", host, path, "", ""))
    key = f"{host}{path}".lower()
    return canonical, key, host


def split_page_parts(path: str) -> tuple[str, str]:
    path = clean_path(path)
    if path == "/":
        return "/", ""
    parts = [part for part in path.split("/") if part]
    page = f"/{parts[0]}"
    sub_page = "/" + "/".join(parts[1:]) if len(parts) > 1 else ""
    return page, sub_page


def is_webpage(url: str) -> bool:
    parsed = urlsplit(url)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        return False
    if parsed.query:
        return False
    path = clean_path(parsed.path).lower()
    if any(path.endswith(ext) for ext in SKIP_EXTENSIONS):
        return False
    return True


def should_exclude_path(path: str, exclude_paths: list[str]) -> bool:
    path = clean_path(path)
    for raw_excluded_path in exclude_paths:
        excluded_path = clean_path(str(raw_excluded_path))
        if excluded_path == "/":
            if path == "/":
                return True
            continue
        if path == excluded_path or path.startswith(f"{excluded_path}/"):
            return True
    return False


def fetch_xml(url: str) -> ElementTree.Element:
    response = requests.get(url, timeout=30, headers={"User-Agent": "website-content-scan/1.0"})
    response.raise_for_status()
    return ElementTree.fromstring(response.content)


def read_xml_file(path: Path) -> ElementTree.Element:
    if path.suffix == ".gz":
        with gzip.open(path, "rb") as handle:
            return ElementTree.fromstring(handle.read())
    return ElementTree.parse(path).getroot()


def xml_name(element: ElementTree.Element) -> str:
    return element.tag.rsplit("}", 1)[-1]


def child_text(element: ElementTree.Element, name: str) -> str:
    for child in element:
        if xml_name(child) == name and child.text:
            return child.text.strip()
    return ""


def iter_sitemap_urls(sitemap_url: str, seen: set[str] | None = None) -> list[tuple[str, str]]:
    seen = seen or set()
    if sitemap_url in seen:
        return []
    seen.add(sitemap_url)

    root = fetch_xml(sitemap_url)
    return iter_sitemap_root(root, sitemap_url, seen)


def iter_sitemap_file(path: Path, seen: set[str] | None = None) -> list[tuple[str, str]]:
    seen = seen or set()
    key = str(path.resolve())
    if key in seen:
        return []
    seen.add(key)

    root = read_xml_file(path)
    return iter_sitemap_root(root, str(path), seen, path.parent)


def iter_sitemap_root(
    root: ElementTree.Element,
    source: str,
    seen: set[str],
    local_base_dir: Path | None = None,
) -> list[tuple[str, str]]:
    root_name = xml_name(root)
    entries: list[tuple[str, str]] = []

    if root_name == "sitemapindex":
        for sitemap in root:
            loc = child_text(sitemap, "loc")
            if not loc:
                continue
            if urlsplit(loc).scheme in {"http", "https"}:
                entries.extend(iter_sitemap_urls(loc, seen))
            elif local_base_dir:
                entries.extend(iter_sitemap_file(local_base_dir / loc, seen))
        return entries

    if root_name != "urlset":
        raise ValueError(f"Unsupported sitemap root '{root_name}' at {source}")

    for url_node in root:
        loc = child_text(url_node, "loc")
        if not loc or not is_webpage(loc):
            continue
        entries.append((loc, child_text(url_node, "lastmod")))
    return entries


def sitemap_pages(config_path: Path) -> dict[str, SitemapPage]:
    pages: dict[str, SitemapPage] = {}
    for site in load_config(config_path):
        exclude_paths = site.get("exclude_paths", [])
        try:
            sitemap_entries = iter_sitemap_urls(site["sitemap"])
        except Exception as exc:
            fallback_file = site.get("fallback_sitemap_file")
            if fallback_file:
                fallback_path = (config_path.parent / fallback_file).resolve()
                if fallback_path.exists():
                    print(
                        f"Warning: using manual sitemap for {site['name']} because {site['sitemap']} could not be read: {exc}",
                        file=sys.stderr,
                    )
                    sitemap_entries = iter_sitemap_file(fallback_path)
                elif site.get("allow_sitemap_errors"):
                    print(
                        f"Warning: skipped {site['name']} because {site['sitemap']} could not be read and {fallback_file} was not found: {exc}",
                        file=sys.stderr,
                    )
                    continue
                else:
                    raise
            elif site.get("allow_sitemap_errors"):
                print(f"Warning: skipped {site['name']} because {site['sitemap']} could not be read: {exc}", file=sys.stderr)
                continue
            else:
                raise
        for raw_url, lastmod in sitemap_entries:
            canonical, key, host = normalize_url(raw_url)
            path = clean_path(urlsplit(canonical).path)
            if should_exclude_path(path, exclude_paths):
                continue
            page, sub_page = split_page_parts(path)
            pages[key] = SitemapPage(
                url=canonical,
                key=key,
                site=host,
                page=page,
                sub_page=sub_page,
                last_updated=lastmod[:10] if lastmod else "",
            )
    return pages


def header_map(ws: Any) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column
    return headers


def ensure_excel_columns(ws: Any) -> dict[str, int]:
    headers = header_map(ws)
    for name in REQUIRED_COLUMNS:
        if name not in headers:
            ws.cell(row=1, column=ws.max_column + 1, value=name)
            headers = header_map(ws)
    return headers


def row_url_key(row: int, ws: Any, headers: dict[str, int]) -> str:
    url = ws.cell(row=row, column=headers["URL"]).value
    if url:
        try:
            return normalize_url(str(url))[1]
        except ValueError:
            pass

    site = ws.cell(row=row, column=headers["Site"]).value
    page = ws.cell(row=row, column=headers["Page"]).value or "/"
    sub_page = ws.cell(row=row, column=headers["Sub-Page"]).value or ""
    if not site:
        return ""
    path = clean_path(f"{page}{sub_page}")
    return f"{strip_www(str(site))}{path}".lower()


def update_excel(input_path: Path, output_path: Path, pages: dict[str, SitemapPage], expired_status: str) -> dict[str, int]:
    workbook = openpyxl.load_workbook(input_path)
    ws = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
    headers = ensure_excel_columns(ws)

    existing: dict[str, int] = {}
    for row in range(2, ws.max_row + 1):
        key = row_url_key(row, ws, headers)
        if key:
            existing[key] = row

    added = updated = expired = 0
    for key, page in sorted(pages.items(), key=lambda item: item[0]):
        if key in existing:
            ws.cell(row=existing[key], column=headers["Last Updated"], value=page.last_updated or None)
            updated += 1
            continue
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=headers["URL"], value=page.url)
        ws.cell(row=new_row, column=headers["Site"], value=page.site)
        ws.cell(row=new_row, column=headers["Page"], value=page.page)
        ws.cell(row=new_row, column=headers["Sub-Page"], value=page.sub_page or None)
        ws.cell(row=new_row, column=headers["Last Updated"], value=page.last_updated or None)
        ws.cell(row=new_row, column=headers["Page Status"], value=DEFAULT_NEW_STATUS)
        added += 1

    for key, row in existing.items():
        if key not in pages:
            ws.cell(row=row, column=headers["Page Status"], value=expired_status)
            expired += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return {"added": added, "updated_last_updated": updated, "expired": expired}


def smartsheet_headers(sheet: dict[str, Any]) -> dict[str, int]:
    return {column["title"]: column["id"] for column in sheet["columns"]}


def smartsheet_request(method: str, url: str, token: str, payload: Any | None = None) -> Any:
    response = requests.request(
        method,
        url,
        timeout=30,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        },
        json=payload,
    )
    response.raise_for_status()
    return response.json() if response.content else {}


def smartsheet_cell(row: dict[str, Any], column_id: int) -> Any:
    for cell in row.get("cells", []):
        if cell.get("columnId") == column_id:
            return cell.get("value")
    return None


def update_smartsheet(sheet_id: str, pages: dict[str, SitemapPage], expired_status: str) -> dict[str, int]:
    token = os.environ.get("SMARTSHEET_ACCESS_TOKEN")
    if not token:
        raise ValueError("SMARTSHEET_ACCESS_TOKEN is required for Smartsheet mode")

    base_url = "https://api.smartsheet.com/2.0"
    sheet = smartsheet_request("GET", f"{base_url}/sheets/{sheet_id}", token)
    columns = smartsheet_headers(sheet)
    missing = [name for name in REQUIRED_COLUMNS if name not in columns]
    if missing:
        raise ValueError(f"Missing Smartsheet columns: {', '.join(missing)}")

    existing: dict[str, dict[str, Any]] = {}
    for row in sheet.get("rows", []):
        url = smartsheet_cell(row, columns["URL"])
        if url:
            key = normalize_url(str(url))[1]
        else:
            site = smartsheet_cell(row, columns["Site"])
            page = smartsheet_cell(row, columns["Page"]) or "/"
            sub_page = smartsheet_cell(row, columns["Sub-Page"]) or ""
            if not site:
                continue
            key = f"{strip_www(str(site))}{clean_path(f'{page}{sub_page}')}".lower()
        existing[key] = row

    rows_to_add = []
    rows_to_update = []

    for key, page in sorted(pages.items(), key=lambda item: item[0]):
        if key in existing:
            rows_to_update.append(
                {
                    "id": existing[key]["id"],
                    "cells": [{"columnId": columns["Last Updated"], "value": page.last_updated or ""}],
                }
            )
        else:
            rows_to_add.append(
                {
                    "toBottom": True,
                    "cells": [
                        {"columnId": columns["Site"], "value": page.site},
                        {"columnId": columns["Page"], "value": page.page},
                        {"columnId": columns["Sub-Page"], "value": page.sub_page},
                        {"columnId": columns["Last Updated"], "value": page.last_updated or ""},
                        {"columnId": columns["Page Status"], "value": DEFAULT_NEW_STATUS},
                    ],
                }
            )

    for key, row in existing.items():
        if key not in pages:
            rows_to_update.append(
                {
                    "id": row["id"],
                    "cells": [{"columnId": columns["Page Status"], "value": expired_status}],
                }
            )

    if rows_to_add:
        smartsheet_request("POST", f"{base_url}/sheets/{sheet_id}/rows", token, rows_to_add)
    for chunk_start in range(0, len(rows_to_update), 500):
        chunk = rows_to_update[chunk_start : chunk_start + 500]
        if chunk:
            smartsheet_request("PUT", f"{base_url}/sheets/{sheet_id}/rows", token, chunk)

    return {
        "added": len(rows_to_add),
        "updated_last_updated": len(pages) - len(rows_to_add),
        "expired": sum(1 for key in existing if key not in pages),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG)
    parser.add_argument("--expired-status", default=DEFAULT_EXPIRED_STATUS)
    parser.add_argument("--excel-input", type=Path, help="Excel export to update as a fallback workflow")
    parser.add_argument("--excel-output", type=Path, help="Where to write the updated Excel workbook")
    parser.add_argument("--smartsheet-sheet-id", default=os.environ.get("SMARTSHEET_SHEET_ID"))
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    pages = sitemap_pages(args.config)

    if args.excel_input:
        output = args.excel_output or args.excel_input.with_name(f"{args.excel_input.stem}.updated.xlsx")
        summary = update_excel(args.excel_input, output, pages, args.expired_status)
        summary["excel_output"] = str(output)
    elif args.smartsheet_sheet_id:
        summary = update_smartsheet(args.smartsheet_sheet_id, pages, args.expired_status)
    else:
        raise SystemExit("Provide --excel-input or SMARTSHEET_SHEET_ID/--smartsheet-sheet-id")

    summary["sitemap_pages_found"] = len(pages)
    summary["finished_at"] = dt.datetime.now(dt.timezone.utc).isoformat()
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
